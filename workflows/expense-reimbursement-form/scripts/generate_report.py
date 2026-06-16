#!/usr/bin/env python3
"""Generate an Expense Reimbursement form by filling the .xlsx template.

Fills only the *input* cells; the template's own formulas (per-row VAT split and
the column totals) are left intact and recalculate when the workbook is opened in
Excel/LibreOffice.

The template models VAT in reverse: column H is the gross amount **incl. VAT**
(in NOK), column I is the Norwegian VAT *rate*, and the template derives the VAT
amount (J = H*I/(1+I)) and the net (K = H-J). Foreign-currency purchases carry no
deductible Norwegian VAT, so their rate is 0.

Usage:
    python3 scripts/generate_report.py \
        --report   data/<batch>/report.json \
        --profile  profile.json \
        --template templates/expense-reimbursement-report.xlsx \
        --out      data/<batch>/output/<batch>.xlsx

See report.example.json for the input schema and .claude/rules/04-output-excel.md.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl

SHEET = "Expense Reimbursement"

# Categories — must match the form's intended dropdown list (10 values).
CATEGORIES = {
    "Transport",
    "Hotel",
    "Fuel",
    "Meals",
    "Entertainment",
    "Phone",
    "Office supplies",
    "Parking/Toll",
    "Postage/Courier",
    "Misc",
}

# Allowed Norwegian VAT rates (column I). 0 = foreign / VAT-exempt.
VAT_RATES = (0.25, 0.15, 0.12, 0.0)
RATE_TOL = 1e-6

# Ledger data area (inclusive): rows 13–106 = 94 rows, above the totals at 107.
LEDGER_FIRST = 13
LEDGER_LAST = 106

# Header input cells.
CELL_COMPANY = "C4"
CELL_CLAIM_DATE = "C5"
CELL_NAME = "C9"
CELL_BANK = "C10"

# Total cells (template formulas) used for the cross-check.
TOTAL_GROSS = "H107"
TOTAL_VAT = "J107"
TOTAL_NET = "K107"


def die(msg: str) -> None:
    print(f"error: {msg}", file=sys.stderr)
    raise SystemExit(1)


def _find_soffice():
    """Locate the LibreOffice CLI, or None."""
    for c in (
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        shutil.which("soffice"),
        shutil.which("libreoffice"),
    ):
        if c and os.path.exists(c):
            return c
    return None


def bake_values(path: Path) -> bool:
    """Recalculate the workbook in place so computed totals are stored as cached
    values. openpyxl writes formulas but no results, so without this some viewers
    (Numbers, Excel for Mac) show blank/zero totals until a manual recalc. Uses
    LibreOffice headless; returns False (with a warning) if it isn't available —
    the file still carries fullCalcOnLoad so Excel will recalc on open."""
    soffice = _find_soffice()
    if not soffice:
        print(
            "warning: LibreOffice not found — totals will recalc on open but are "
            "not pre-cached. Install LibreOffice for baked values.",
            file=sys.stderr,
        )
        return False
    tmp = path.parent / "_recalc"
    tmp.mkdir(exist_ok=True)
    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                "-env:UserInstallation=file:///tmp/lo_recalc_profile",
                "--calc",
                "--convert-to",
                "xlsx:Calc MS Excel 2007 XML",
                "--outdir",
                str(tmp),
                str(path),
            ],
            check=True,
            capture_output=True,
            timeout=180,
        )
        produced = tmp / f"{path.stem}.xlsx"
        if produced.exists():
            shutil.move(str(produced), str(path))
            return True
        print(
            "warning: recalc produced no file; totals will recalc on open.",
            file=sys.stderr,
        )
        return False
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        print(
            f"warning: recalc failed ({e}); totals will recalc on open.",
            file=sys.stderr,
        )
        return False
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _cellnum(ws, coord) -> float:
    v = ws[coord].value
    return float(v) if v is not None else 0.0


def validate_totals(path: Path, txns: list) -> None:
    """Cross-check the template's computed totals against sums taken directly from
    the transactions: gross (incl. VAT), the extracted Norwegian VAT, and net
    (excl. VAT). Catches a row that landed outside a SUM range or a rate that did
    not write through. Requires the workbook to have been recalculated (baked);
    skips with a warning otherwise."""
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    if ws[TOTAL_GROSS].value is None:  # no cached results -> can't compare
        print(
            "warning: totals not pre-computed (no recalc) — skipping totals "
            "cross-check. Install LibreOffice to enable it.",
            file=sys.stderr,
        )
        return

    TOL = 0.01
    gross = sum(float(t["amount_incl_vat"]) for t in txns)
    vat = sum(
        float(t["amount_incl_vat"]) * float(t["vat_rate"]) / (1 + float(t["vat_rate"]))
        for t in txns
    )
    net = gross - vat

    problems = []
    for coord, exp, label in (
        (TOTAL_GROSS, gross, "gross incl. VAT"),
        (TOTAL_VAT, vat, "Norwegian VAT"),
        (TOTAL_NET, net, "net excl. VAT"),
    ):
        got = _cellnum(ws, coord)
        if abs(got - exp) > TOL:
            problems.append(f"{label} {coord}={got:.2f}, sum of transactions={exp:.2f}")

    if problems:
        die(
            "computed totals do not match the transactions:\n  - "
            + "\n  - ".join(problems)
        )
    print(
        f"  totals validated: {gross:.2f} NOK gross, {vat:.2f} VAT, {net:.2f} net "
        f"== sum of {len(txns)} transactions"
    )


def parse_date(s):
    """ISO string -> date (no time). Accepts a date already."""
    if s is None or s == "":
        return None
    if isinstance(s, dt.date):
        return s
    return dt.date.fromisoformat(str(s).strip())


def num(v):
    """Return a float for a present numeric value, or None to leave the cell blank."""
    if v is None or v == "":
        return None
    return float(v)


def parse_rate(v):
    """Normalise a VAT rate to a decimal fraction and validate it.

    Accepts a fraction (0.25) or a percent (25 -> 0.25). Must resolve to one of
    the template's allowed rates."""
    if v is None or v == "":
        die("vat_rate is mandatory (use 0 for foreign / VAT-exempt rows).")
    r = float(v)
    if r > 1:  # given as a percent, e.g. 25
        r = r / 100.0
    for allowed in VAT_RATES:
        if abs(r - allowed) < RATE_TOL:
            return allowed
    die(f"vat_rate {v!r} -> {r} is not one of {sorted(VAT_RATES, reverse=True)}.")


def fill(report: dict, profile: dict, template: Path, out: Path) -> None:
    wb = openpyxl.load_workbook(template)  # keep formulas (data_only=False)
    ws = wb[SHEET]

    # ── Header ────────────────────────────────────────────────────────────────
    ws[CELL_COMPANY] = report.get("company") or profile.get("company")
    claim = report.get("claim_date") or dt.date.today().isoformat()
    ws[CELL_CLAIM_DATE] = parse_date(claim)
    ws[CELL_NAME] = profile.get("name")
    ws[CELL_BANK] = profile.get("bank_account")

    # ── Expense ledger ────────────────────────────────────────────────────────
    txns = report.get("transactions", [])
    capacity = LEDGER_LAST - LEDGER_FIRST + 1
    if len(txns) > capacity:
        die(
            f"{len(txns)} transactions exceed the template's {capacity} ledger rows "
            f"(rows {LEDGER_FIRST}-{LEDGER_LAST}). Split the batch."
        )

    for i, t in enumerate(txns):
        cat = t.get("category")
        if cat not in CATEGORIES:
            die(
                f"transaction {i + 1}: category {cat!r} is not one of "
                f"{sorted(CATEGORIES)}."
            )
        cur = t.get("currency")
        if not cur:
            die(f"transaction {i + 1}: currency is mandatory.")
        amount = num(t.get("amount_incl_vat"))
        if amount is None:
            die(f"transaction {i + 1}: amount_incl_vat is mandatory.")
        rate = parse_rate(t.get("vat_rate"))
        if cur != "NOK" and rate != 0.0:
            die(
                f"transaction {i + 1}: foreign-currency ({cur}) rows carry no "
                f"deductible Norwegian VAT — vat_rate must be 0 (got {rate})."
            )

        r = LEDGER_FIRST + i
        ws[f"B{r}"] = parse_date(t.get("date"))
        ws[f"C{r}"] = t.get("attachment_no")  # Bilagsnr, e.g. "01" or "01, 02"
        ws[f"D{r}"] = t.get("supplier")
        ws[f"E{r}"] = t.get("description")
        ws[f"F{r}"] = cat
        ws[f"G{r}"] = cur
        ws[f"H{r}"] = amount
        ws[f"I{r}"] = rate

    # Force a full recalc on open as a fallback for viewers that honor it.
    wb.calculation.fullCalcOnLoad = True

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    # Bake computed totals in so they display in every viewer, not just on recalc.
    baked = bake_values(out)
    if baked:
        validate_totals(out, txns)
    print(
        f"wrote {out}  ({len(txns)} transactions){'  [totals baked]' if baked else ''}"
    )


def main() -> None:
    ap = argparse.ArgumentParser(description="Fill the expense reimbursement template.")
    ap.add_argument("--report", required=True, type=Path)
    ap.add_argument("--profile", required=True, type=Path)
    ap.add_argument("--template", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    args = ap.parse_args()

    for p in (args.report, args.profile, args.template):
        if not p.exists():
            die(f"not found: {p}")

    report = json.loads(args.report.read_text())
    profile = json.loads(args.profile.read_text())
    fill(report, profile, args.template, args.out)


if __name__ == "__main__":
    main()
