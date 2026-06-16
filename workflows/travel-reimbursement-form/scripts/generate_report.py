#!/usr/bin/env python3
"""Generate a Travel Expense Report by filling the .xlsx template.

Fills only the *input* cells; the template's own formulas (allowance totals,
markup, and the bookkeeping summary) are left intact and recalculate when the
workbook is opened in Excel/LibreOffice.

Usage:
    python3 scripts/generate_report.py \
        --report   data/<trip>/report.json \
        --profile  profile.json \
        --template templates/travel-expense-report.xlsx \
        --out      data/<trip>/output/<trip>.xlsx

See report.example.json for the input schema and .claude/rules/06-output-excel.md.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl

# Ledger categories — must match the template's Sheet2 dropdown (7 values).
# "Subsistence and supplements" is NOT a ledger category here: it is the
# allowance section, summed into the markup table by the template.
LEDGER_CATEGORIES = {
    "Hotel",
    "Transport",
    "Fuel",
    "Meals",
    "Phone",
    "Entertainment",
    "Misc",
}

# Ledger data area (inclusive). Stays within the template's validated range and
# clear of the allowance section that starts at row 47.
LEDGER_FIRST = 15
LEDGER_LAST = 42  # 28 rows

# Allowance quantities are written here; rates/totals are template formulas.
ALLOWANCE_FIRST = 47
ALLOWANCE_LAST = 67

# Markup rate cell per category (column H of the markup roll-up).
MARKUP_ROW = {
    "Hotel": 71,
    "Transport": 72,
    "Fuel": 73,
    "Meals": 74,
    "Phone": 75,
    "Entertainment": 76,
    "Misc": 77,
    "Subsistence and supplements": 78,
}


def die(msg: str) -> None:
    print(f"error: {msg}", file=sys.stderr)
    raise SystemExit(1)


def _find_soffice():
    """Locate the LibreOffice CLI, or None."""
    for c in (
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        shutil.which("soffice"),
        shutil.which("libreoffice"),
    ):
        if c and os.path.exists(c):
            return c
    return None


def bake_values(path: Path) -> bool:
    """Recalculate the workbook in place so computed totals are stored as cached
    values. openpyxl writes formulas but no results, so without this some viewers
    (Numbers, Excel for Mac) show blank/zero totals until a manual recalc. Uses
    LibreOffice headless; returns False (with a warning) if it isn't available —
    the file still carries fullCalcOnLoad so Excel will recalc on open."""
    soffice = _find_soffice()
    if not soffice:
        print(
            "warning: LibreOffice not found — totals will recalc on open but are "
            "not pre-cached. Install LibreOffice for baked values.",
            file=sys.stderr,
        )
        return False
    tmp = path.parent / "_recalc"
    tmp.mkdir(exist_ok=True)
    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                "-env:UserInstallation=file:///tmp/lo_recalc_profile",
                "--calc",
                "--convert-to",
                "xlsx:Calc MS Excel 2007 XML",
                "--outdir",
                str(tmp),
                str(path),
            ],
            check=True,
            capture_output=True,
            timeout=180,
        )
        produced = tmp / f"{path.stem}.xlsx"
        if produced.exists():
            shutil.move(str(produced), str(path))
            return True
        print(
            "warning: recalc produced no file; totals will recalc on open.",
            file=sys.stderr,
        )
        return False
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        print(
            f"warning: recalc failed ({e}); totals will recalc on open.",
            file=sys.stderr,
        )
        return False
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _cellnum(ws, coord) -> float:
    v = ws[coord].value
    return float(v) if v is not None else 0.0


def validate_totals(path: Path, txns: list) -> None:
    """Cross-check the template's computed totals against sums taken directly from
    the transactions: the grand total, the markup combined total, and every
    per-category bookkeeping cell (NOK/other x refunded/not). Catches a misplaced
    row, a formula whose range misses data, or a category mismatch. Requires the
    workbook to have been recalculated (baked); skips with a warning otherwise."""
    ws = openpyxl.load_workbook(path, data_only=True)["Sheet1"]
    if ws["G91"].value is None:  # no cached results -> can't compare
        print(
            "warning: totals not pre-computed (no recalc) — skipping totals "
            "cross-check. Install LibreOffice to enable it.",
            file=sys.stderr,
        )
        return

    TOL = 0.01

    def amt(t):
        return float(t.get("amount_nok") or 0)

    def rebilled(t):
        v = t.get("rebilled_nok")
        return v is not None and float(v) > 0

    problems = []
    expected_grand = round(sum(amt(t) for t in txns), 2)
    if abs(_cellnum(ws, "G91") - expected_grand) > TOL:
        problems.append(
            f"grand total G91={_cellnum(ws, 'G91'):.2f}, "
            f"sum of transactions={expected_grand:.2f}"
        )
    if abs(_cellnum(ws, "J79") - expected_grand) > TOL:
        problems.append(
            f"markup combined J79={_cellnum(ws, 'J79'):.2f}, "
            f"expected {expected_grand:.2f}"
        )

    # Per-category bookkeeping rows (category label is in column A).
    for r in range(84, 91):
        cat = ws[f"A{r}"].value
        if not cat:
            continue
        rows = [t for t in txns if t.get("category") == cat]
        nok = sum(amt(t) for t in rows if t.get("original_currency") == "NOK")
        oth = sum(amt(t) for t in rows if t.get("original_currency") != "NOK")
        ref_nok = sum(
            amt(t) for t in rows if t.get("original_currency") == "NOK" and rebilled(t)
        )
        ref_oth = sum(
            amt(t) for t in rows if t.get("original_currency") != "NOK" and rebilled(t)
        )
        expected = {
            "C": round(nok - ref_nok, 2),  # not refunded, NOK
            "D": round(oth - ref_oth, 2),  # not refunded, other currencies
            "E": round(ref_nok, 2),  # refunded, NOK
            "F": round(ref_oth, 2),  # refunded, other currencies
            "G": round(nok + oth, 2),  # category total
        }
        for col, exp in expected.items():
            got = _cellnum(ws, f"{col}{r}")
            if abs(got - exp) > TOL:
                problems.append(f"{cat} {col}{r}={got:.2f}, expected {exp:.2f}")

    if problems:
        die(
            "computed totals do not match the transactions:\n  - "
            + "\n  - ".join(problems)
        )
    print(
        f"  totals validated: {expected_grand:.2f} NOK == sum of {len(txns)} transactions"
    )


def parse_date(s: str):
    """ISO string -> date (no time) or datetime (with time)."""
    if s is None:
        return None
    s = s.strip()
    if "T" in s or " " in s:
        return dt.datetime.fromisoformat(s.replace(" ", "T"))
    return dt.date.fromisoformat(s)


def num(v):
    """Return a float for a present numeric value, or None to leave the cell blank."""
    if v is None or v == "":
        return None
    return float(v)


def fill(report: dict, profile: dict, template: Path, out: Path) -> None:
    wb = openpyxl.load_workbook(template)  # keep formulas (data_only=False)
    ws = wb["Sheet1"]

    # ── Header ────────────────────────────────────────────────────────────────
    ws["B3"] = profile.get("name")
    ws["B4"] = profile.get("address")
    ws["B5"] = profile.get("postal_city")
    ws["B6"] = profile.get("bank_account")
    ws["B7"] = profile.get("email")

    ws["B9"] = parse_date(report.get("travel_start"))
    ws["B10"] = parse_date(report.get("travel_end"))
    submitted = report.get("date_submitted") or dt.date.today().isoformat()
    ws["B11"] = parse_date(submitted)

    # ── Actual expenses ledger ────────────────────────────────────────────────
    txns = report.get("transactions", [])
    capacity = LEDGER_LAST - LEDGER_FIRST + 1
    if len(txns) > capacity:
        die(
            f"{len(txns)} transactions exceed the template's {capacity} ledger rows "
            f"(rows {LEDGER_FIRST}-{LEDGER_LAST}). Extend the template or split the trip."
        )

    for i, t in enumerate(txns):
        cat = t.get("category")
        if cat not in LEDGER_CATEGORIES:
            die(
                f"transaction {i + 1}: category {cat!r} is not a ledger category "
                f"{sorted(LEDGER_CATEGORIES)}. (Subsistence/allowance goes in the "
                f"allowance section, not the ledger.)"
            )
        cur = t.get("original_currency")
        if not cur:
            die(f"transaction {i + 1}: original_currency is mandatory.")
        vat = num(t.get("vat_nok"))
        if cur == "NOK" and vat is None:
            die(f"transaction {i + 1}: vat_nok is required for NOK-currency rows.")
        if cur != "NOK" and vat is not None:
            die(
                f"transaction {i + 1}: vat_nok must be blank for non-NOK rows (got {vat})."
            )

        r = LEDGER_FIRST + i
        ws[f"A{r}"] = parse_date(t.get("date"))
        ws[f"B{r}"] = t.get("description")
        ws[f"F{r}"] = t.get("attachment_no")  # text, e.g. "29, 30, 33"
        ws[f"G{r}"] = num(t.get("amount_nok"))
        ws[f"H{r}"] = vat
        ws[f"I{r}"] = cat
        ws[f"J{r}"] = num(t.get("rebilled_nok"))
        ws[f"K{r}"] = cur

    # ── Allowance quantities (rates/totals stay as template formulas) ─────────
    for entry in report.get("allowance", []):
        row = int(entry["row"])
        if not (ALLOWANCE_FIRST <= row <= ALLOWANCE_LAST):
            die(f"allowance row {row} out of range {ALLOWANCE_FIRST}-{ALLOWANCE_LAST}.")
        q = num(entry.get("quantity"))
        if q is not None:
            ws[f"H{row}"] = q
        qr = num(entry.get("quantity_rebilled"))
        if qr is not None:
            ws[f"I{row}"] = qr

    # ── Markup rates (default left at the template's 0) ───────────────────────
    for cat, rate in (report.get("markup") or {}).items():
        if cat not in MARKUP_ROW:
            die(
                f"markup category {cat!r} unknown; expected one of {sorted(MARKUP_ROW)}."
            )
        ws[f"H{MARKUP_ROW[cat]}"] = float(rate)

    # Force a full recalc on open as a fallback for viewers that honor it.
    wb.calculation.fullCalcOnLoad = True

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    # Bake computed totals in so they display in every viewer, not just on recalc.
    baked = bake_values(out)
    if baked:
        # Guard: the template's computed totals must equal the transaction sums.
        validate_totals(out, txns)
    print(
        f"wrote {out}  ({len(txns)} transactions){'  [totals baked]' if baked else ''}"
    )


def main() -> None:
    ap = argparse.ArgumentParser(description="Fill the travel expense report template.")
    ap.add_argument("--report", required=True, type=Path)
    ap.add_argument("--profile", required=True, type=Path)
    ap.add_argument("--template", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    args = ap.parse_args()

    for p in (args.report, args.profile, args.template):
        if not p.exists():
            die(f"not found: {p}")

    report = json.loads(args.report.read_text())
    profile = json.loads(args.profile.read_text())
    fill(report, profile, args.template, args.out)


if __name__ == "__main__":
    main()
